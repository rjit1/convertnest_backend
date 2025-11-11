"""
Flask API for Image-to-Excel Conversion using Gemini 2.5 Flash
Production-ready service for handwritten table extraction
"""
import os
import sys
import logging
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import traceback
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Lazy import flag - import libraries only when needed
_modules_loaded = False
_tts_loaded = False
GeminiTableExtractor = None
ExcelGenerator = None
PDFToExcelExtractor = None
UsageTracker = None
GeminiTTSService = None

def _load_modules():
    """Lazy load Gemini modules only when first needed"""
    global _modules_loaded, GeminiTableExtractor, ExcelGenerator, PDFToExcelExtractor, UsageTracker
    if not _modules_loaded:
        logger.info("Loading Gemini Table Extractor...")
        from gemini_table_extractor import GeminiTableExtractor as GTE
        from excel_generator import ExcelGenerator as EG
        from pdf_to_excel_extractor import PDFToExcelExtractor as PTE
        from usage_tracker import get_usage_tracker
        GeminiTableExtractor = GTE
        ExcelGenerator = EG
        PDFToExcelExtractor = PTE
        UsageTracker = get_usage_tracker
        _modules_loaded = True
        logger.info("✓ Gemini modules loaded successfully")

def _load_tts():
    """Lazy load TTS module separately to avoid PDF dependencies"""
    global _tts_loaded, GeminiTTSService
    if not _tts_loaded:
        logger.info("Loading Gemini TTS Service...")
        from gemini_tts_service import GeminiTTSService as TTS
        GeminiTTSService = TTS
        _tts_loaded = True
        logger.info("✓ TTS module loaded successfully")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
OUTPUT_FOLDER = os.path.join(os.path.dirname(__file__), 'outputs')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'bmp', 'tiff', 'webp'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Create directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Initialize services (lazy loading to speed up startup)
table_extractor = None
excel_generator = None
pdf_to_excel_extractor = None
tts_service = None

def get_table_extractor():
    """Lazy initialization of Gemini table extractor"""
    global table_extractor
    if table_extractor is None:
        _load_modules()  # Load modules if not already loaded
        logger.info("Initializing Gemini Table Extractor...")
        table_extractor = GeminiTableExtractor()
        logger.info("✓ Gemini Table Extractor initialized")
    return table_extractor

def get_excel_generator():
    """Lazy initialization of excel generator"""
    global excel_generator
    if excel_generator is None:
        _load_modules()  # Load modules if not already loaded
        logger.info("Initializing Excel Generator...")
        excel_generator = ExcelGenerator()
        logger.info("✓ Excel Generator initialized")
    return excel_generator

def get_pdf_to_excel_extractor():
    """Lazy initialization of PDF-to-Excel extractor"""
    global pdf_to_excel_extractor
    if pdf_to_excel_extractor is None:
        _load_modules()  # Load modules if not already loaded
        logger.info("Initializing PDF-to-Excel Extractor...")
        pdf_to_excel_extractor = PDFToExcelExtractor()
        logger.info("✓ PDF-to-Excel Extractor initialized")
    return pdf_to_excel_extractor

def get_tts_service():
    """Lazy initialization of TTS service"""
    global tts_service
    if tts_service is None:
        _load_tts()  # Load TTS module separately (no PDF dependencies)
        logger.info("Initializing Gemini TTS Service...")
        tts_service = GeminiTTSService()
        logger.info("✓ Gemini TTS Service initialized")
    return tts_service

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'Gemini Conversion Service',
        'version': '2.0.0',
        'features': {
            'image_to_excel': 'gemini-2.5-flash',
            'pdf_to_excel': 'gemini-2.5-flash-sequential (max 25 pages)',
            'text_to_speech': 'gemini-2.5-pro-preview-tts (30 voices, 24 languages)'
        },
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/extract-table', methods=['POST'])
def extract_table():
    """
    Main endpoint for table extraction using Gemini
    
    Expects:
    - file: Image file containing table
    
    Returns:
    - Excel file with extracted table
    """
    try:
        # Validate request
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Empty filename'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': f'Invalid file type. Allowed: {ALLOWED_EXTENSIONS}'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        input_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(input_path)
        
        logger.info(f"Processing file: {unique_filename}")
        
        # Extract table using Gemini
        extractor = get_table_extractor()
        extraction_result = extractor.extract_tables(
            image_path=input_path
        )
        
        if not extraction_result or not extraction_result.get('tables'):
            logger.warning(f"No tables detected in {unique_filename}")
            # Clean up
            if os.path.exists(input_path):
                os.remove(input_path)
            return jsonify({
                'error': 'No tables detected in image',
                'suggestion': 'Please ensure the image contains a clear table with visible structure'
            }), 404
        
        logger.info(f"Detected {len(extraction_result['tables'])} table(s)")
        
        # Generate Excel file
        generator = get_excel_generator()
        output_filename = f"{timestamp}_table_extracted.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)
        
        excel_result = generator.create_excel(
            tables=extraction_result['tables'],
            output_dir=OUTPUT_FOLDER,
            filename=output_filename
        )
        
        if not excel_result.get('success'):
            raise Exception(excel_result.get('error', 'Failed to create Excel file'))
        
        logger.info(f"Excel file created: {output_filename} ({excel_result['file_size']} bytes)")
        
        # Clean up input file
        if os.path.exists(input_path):
            os.remove(input_path)
        
        # Send Excel file
        return send_file(
            excel_result['excel_path'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"table_extracted_{filename.rsplit('.', 1)[0]}.xlsx"
        )
        
    except Exception as e:
        logger.error(f"Error processing request: {str(e)}")
        logger.error(traceback.format_exc())
        
        # Clean up files on error
        if 'input_path' in locals() and os.path.exists(input_path):
            os.remove(input_path)
        if 'output_path' in locals() and os.path.exists(output_path):
            os.remove(output_path)
        
        return jsonify({
            'error': 'Internal server error',
            'message': str(e),
            'type': type(e).__name__
        }), 500

@app.route('/api/extract-table-json', methods=['POST'])
def extract_table_json():
    """
    Extract table and return JSON representation instead of Excel
    
    Useful for debugging and custom processing
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '' or not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file'}), 400
        
        # Save file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        input_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(input_path)
        
        # Extract tables
        extractor = get_table_extractor()
        result = extractor.extract_tables(image_path=input_path)
        
        # Clean up
        if os.path.exists(input_path):
            os.remove(input_path)
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error in JSON extraction: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/pdf-info', methods=['POST'])
def get_pdf_info():
    """
    Get PDF information (page count, validation)
    
    Expects:
    - file: PDF file
    
    Returns:
    - JSON with page count and validation
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '' or not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Invalid file. Only PDF files are allowed.'}), 400
        
        # Save file temporarily
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        input_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{filename}")
        file.save(input_path)
        
        logger.info(f"Getting PDF info: {filename}")
        
        # Get PDF info
        extractor = get_pdf_to_excel_extractor()
        validation = extractor.validate_pdf(input_path)
        
        # Clean up
        if os.path.exists(input_path):
            os.remove(input_path)
        
        return jsonify({
            'success': validation['valid'],
            'page_count': validation['page_count'],
            'file_size': validation.get('file_size', 0),
            'error': validation.get('error')
        })
        
    except Exception as e:
        logger.error(f"Error getting PDF info: {str(e)}")
        
        # Clean up on error
        if 'input_path' in locals() and os.path.exists(input_path):
            os.remove(input_path)
        
        return jsonify({'error': str(e)}), 500

@app.route('/api/pdf-to-excel', methods=['POST'])
def convert_pdf_to_excel():
    """
    Convert multi-page PDF to Excel with one sheet per page (PARALLEL PROCESSING)
    
    Expects:
    - file: PDF file (max 25 pages)
    
    Returns:
    - Excel file with multiple sheets (1 sheet per page)
    
    Features:
    - Parallel Gemini API calls (10x faster than sequential)
    - Daily usage limit: 100 conversions per day
    - Returns quota info in headers
    """
    try:
        # Load modules (including usage tracker)
        _load_modules()
        
        # Step 1: Check daily quota BEFORE processing
        tracker = UsageTracker()
        allowed, remaining, message = tracker.check_quota()
        
        if not allowed:
            usage_info = tracker.get_usage_info()
            logger.warning(f"⛔ Quota exceeded: {message}")
            
            return jsonify({
                'success': False,
                'error': 'Daily Quota Limit Reached',
                'message': f'You have used all {usage_info["limit"]} free conversions for today.',
                'details': {
                    'quota_status': 'exceeded',
                    'used_today': usage_info['used'],
                    'daily_limit': usage_info['limit'],
                    'remaining': 0,
                    'reset_time': 'midnight UTC',
                    'current_date': usage_info['date']
                },
                'user_message': f'Daily limit reached! You\'ve converted {usage_info["used"]} PDFs today (limit: {usage_info["limit"]}). Your quota will reset at midnight UTC. Please come back tomorrow!',
                'suggestions': [
                    'Wait until midnight UTC for quota reset',
                    'Try again tomorrow',
                    'Contact support if you need higher limits'
                ]
            }), 429  # 429 Too Many Requests
        
        logger.info(f"✅ Quota check passed: {remaining} conversions remaining")
        
        # Step 2: Validate request
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '' or not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Invalid file. Only PDF files are allowed.'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        input_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        file.save(input_path)
        
        logger.info(f"Processing PDF to Excel: {unique_filename}")
        
        # Get extractor
        extractor = get_pdf_to_excel_extractor()
        
        # Validate PDF first
        validation = extractor.validate_pdf(input_path)
        if not validation['valid']:
            # Clean up
            if os.path.exists(input_path):
                os.remove(input_path)
            return jsonify({'error': validation['error']}), 400
        
        logger.info(f"PDF validated: {validation['page_count']} pages")
        
        # Process PDF (extract pages and call Gemini for each)
        result = extractor.process_pdf(input_path)
        
        if not result['success']:
            # Clean up
            if os.path.exists(input_path):
                os.remove(input_path)
            return jsonify({
                'error': result.get('error', 'PDF processing failed'),
                'details': {
                    'total_pages': result['total_pages'],
                    'successful_pages': result['successful_pages'],
                    'failed_pages': result['failed_pages']
                }
            }), 500
        
        logger.info(f"PDF processing complete: {result['successful_pages']}/{result['total_pages']} pages")
        
        # Generate Excel file from all page results
        # Convert page results to table format for Excel generator
        tables_for_excel = []
        
        for page_result in result['page_results']:
            if page_result['success'] and page_result.get('tables'):
                # Each page gets its own sheet
                for table in page_result['tables']:
                    # Add page number to table metadata
                    table_with_page = {
                        'table_id': page_result['page_number'],
                        'page_number': page_result['page_number'],
                        'headers': table.get('headers', []),
                        'data': table.get('data', []),
                        'metadata': table.get('metadata', {})
                    }
                    tables_for_excel.append(table_with_page)
            else:
                # Add empty sheet for failed pages
                logger.warning(f"Page {page_result['page_number']} failed: {page_result.get('error', 'Unknown')}")
                tables_for_excel.append({
                    'table_id': page_result['page_number'],
                    'page_number': page_result['page_number'],
                    'headers': ['Error'],
                    'data': [[f"Failed to process page {page_result['page_number']}: {page_result.get('error', 'Unknown error')}"]]
                })
        
        if not tables_for_excel:
            # Clean up
            if os.path.exists(input_path):
                os.remove(input_path)
            extractor.cleanup_temp_files(result.get('temp_dir', ''), result.get('image_paths', []))
            return jsonify({'error': 'No tables extracted from any page'}), 404
        
        # Generate Excel with custom multi-sheet logic
        output_filename = f"{timestamp}_{filename.rsplit('.', 1)[0]}_converted.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)
        
        # Use openpyxl to create multi-sheet Excel
        from openpyxl import Workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create one sheet per page
        for table_data in tables_for_excel:
            page_num = table_data['page_number']
            sheet_name = f"Page {page_num}"
            ws = wb.create_sheet(title=sheet_name)
            
            # Write headers
            headers = table_data.get('headers', [])
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                start_row = 2
            else:
                start_row = 1
            
            # Write data
            data_rows = table_data.get('data', [])
            for row_idx, row_data in enumerate(data_rows, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Save Excel file
        wb.save(output_path)
        file_size = os.path.getsize(output_path)
        
        logger.info(f"Excel file created: {output_filename} ({file_size:,} bytes)")
        logger.info(f"   Sheets: {len(tables_for_excel)} | Processing time: {result['processing_time']:.2f}s")
        
        # Increment usage counter (successful conversion)
        usage_info = tracker.increment_usage()
        logger.info(f"📊 Usage updated: {usage_info['used']}/{usage_info['limit']} ({usage_info['remaining']} remaining)")
        
        # Clean up temporary files
        if os.path.exists(input_path):
            os.remove(input_path)
        extractor.cleanup_temp_files(result.get('temp_dir', ''), result.get('image_paths', []))
        
        # Send Excel file with metadata in headers
        response = send_file(
            output_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename
        )
        
        # Add custom headers with processing info
        response.headers['X-Processing-Time'] = str(int(result['processing_time']))
        response.headers['X-Total-Pages'] = str(result['total_pages'])
        response.headers['X-Successful-Pages'] = str(result['successful_pages'])
        response.headers['X-Failed-Pages'] = str(result['failed_pages'])
        response.headers['X-Processing-Mode'] = 'parallel'  # Indicate parallel processing
        
        # Add quota headers
        response.headers['X-Daily-Quota-Used'] = str(usage_info['used'])
        response.headers['X-Daily-Quota-Limit'] = str(usage_info['limit'])
        response.headers['X-Daily-Quota-Remaining'] = str(usage_info['remaining'])
        
        return response
        
    except Exception as e:
        logger.error(f"Error in PDF to Excel conversion: {str(e)}")
        logger.error(traceback.format_exc())
        
        # Clean up files on error
        if 'input_path' in locals() and os.path.exists(input_path):
            os.remove(input_path)
        if 'output_path' in locals() and os.path.exists(output_path):
            os.remove(output_path)
        if 'result' in locals() and 'temp_dir' in result:
            try:
                extractor.cleanup_temp_files(result.get('temp_dir', ''), result.get('image_paths', []))
            except:
                pass
        
        return jsonify({
            'error': 'Internal server error',
            'message': str(e),
            'type': type(e).__name__
        }), 500

@app.route('/api/pdf-quota', methods=['GET'])
def get_pdf_quota():
    """
    Get current PDF to Excel conversion quota information.
    
    Returns:
        JSON with quota details (used, limit, remaining)
    """
    try:
        _load_modules()
        tracker = UsageTracker()
        usage_info = tracker.get_usage_info()
        
        return jsonify({
            'success': True,
            'quota': {
                'date': usage_info['date'],
                'used': usage_info['used'],
                'limit': usage_info['limit'],
                'remaining': usage_info['remaining'],
                'percentage': round(usage_info['percentage'], 1)
            },
            'service': 'pdf-to-excel',
            'processing_mode': 'parallel',
            'features': {
                'max_pages_per_pdf': 25,
                'estimated_time_20_pages': '4-8 seconds',
                'model': 'gemini-2.5-flash'
            }
        })
    except Exception as e:
        logger.error(f"Error getting quota info: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================================
# TEXT-TO-SPEECH ROUTES (Gemini 2.5 Pro TTS)
# ============================================================================

@app.route('/api/tts/voices', methods=['GET'])
def get_tts_voices():
    """
    Get complete voice catalog with categories
    
    Returns:
        JSON with 30 voices organized by type (professional, friendly, calm, etc.)
    """
    try:
        tts = get_tts_service()
        catalog = tts.get_voices_catalog()
        
        return jsonify({
            'success': True,
            'voices': catalog
        })
    except Exception as e:
        logger.error(f"Error getting voices: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/tts/languages', methods=['GET'])
def get_tts_languages():
    """
    Get supported languages
    
    Returns:
        JSON with 24 supported languages
    """
    try:
        tts = get_tts_service()
        languages = tts.get_languages()
        
        return jsonify({
            'success': True,
            'languages': languages
        })
    except Exception as e:
        logger.error(f"Error getting languages: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/tts/generate', methods=['POST'])
def generate_tts():
    """
    Generate speech from text (single speaker)
    
    Expects JSON:
    {
        "text": "Text to convert",
        "voice": "Kore" (optional, default: Kore),
        "style": "Say cheerfully:" (optional),
        "language": "en-US" (optional, default: en-US),
        "temperature": 0.7 (optional, default: 0.7)
    }
    
    Returns:
        WAV audio file
    """
    try:
        # Get JSON data
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        text = data.get('text', '').strip()
        if not text:
            return jsonify({'error': 'Empty text provided'}), 400
        
        # Extract parameters (HIGH IMPACT ONLY)
        voice = data.get('voice', 'Kore')
        style = data.get('style', '')
        language = data.get('language', 'en-US')
        temperature = float(data.get('temperature', 0.7))
        
        # Validate temperature
        if temperature < 0.0 or temperature > 2.0:
            temperature = 0.7
        
        logger.info(f"🎤 TTS Request: {len(text)} chars, voice={voice}, lang={language}")
        
        # Generate speech
        tts = get_tts_service()
        result = tts.generate_speech(
            text=text,
            voice_name=voice,
            style_prompt=style,
            language_code=language,
            temperature=temperature
        )
        
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Speech generation failed')}), 500
        
        logger.info(f"✓ TTS generated: {result['filename']} ({result['file_size']:,} bytes)")
        
        # Send WAV file
        response = send_file(
            result['audio_path'],
            mimetype='audio/wav',
            as_attachment=True,
            download_name=result['filename']
        )
        
        # Add metadata headers
        response.headers['X-Voice'] = result['voice']
        response.headers['X-Language'] = result['language']
        response.headers['X-Duration'] = str(int(result['duration']))
        response.headers['X-Char-Count'] = str(result['char_count'])
        
        return response
        
    except Exception as e:
        logger.error(f"Error in TTS generation: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'error': 'TTS generation failed',
            'message': str(e),
            'type': type(e).__name__
        }), 500

@app.route('/api/tts/generate-dialog', methods=['POST'])
def generate_tts_dialog():
    """
    Generate multi-speaker dialog (up to 2 speakers)
    
    Expects JSON:
    {
        "speakers": [
            {"name": "Speaker1", "voice": "Kore"},
            {"name": "Speaker2", "voice": "Puck"}
        ],
        "transcript": "Speaker1: Hello! Speaker2: Hi there!",
        "language": "en-US" (optional),
        "temperature": 0.7 (optional)
    }
    
    Returns:
        WAV audio file with dialog
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No JSON data provided'}), 400
        
        speakers = data.get('speakers', [])
        transcript = data.get('transcript', '').strip()
        language = data.get('language', 'en-US')
        temperature = float(data.get('temperature', 0.7))
        
        if not speakers:
            return jsonify({'error': 'No speakers provided'}), 400
        if not transcript:
            return jsonify({'error': 'Empty transcript'}), 400
        
        logger.info(f"🎭 Dialog TTS: {len(speakers)} speakers, {len(transcript)} chars")
        
        # Generate dialog
        tts = get_tts_service()
        result = tts.generate_dialog(
            speakers=speakers,
            transcript=transcript,
            language_code=language,
            temperature=temperature
        )
        
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Dialog generation failed')}), 500
        
        logger.info(f"✓ Dialog generated: {result['filename']} ({result['file_size']:,} bytes)")
        
        # Send WAV file
        response = send_file(
            result['audio_path'],
            mimetype='audio/wav',
            as_attachment=True,
            download_name=result['filename']
        )
        
        # Add metadata headers
        response.headers['X-Speakers'] = ','.join(result['speakers'])
        response.headers['X-Voices'] = ','.join(result['voices'])
        response.headers['X-Language'] = result['language']
        response.headers['X-Duration'] = str(int(result['duration']))
        
        return response
        
    except Exception as e:
        logger.error(f"Error in dialog generation: {str(e)}")
        return jsonify({
            'error': 'Dialog generation failed',
            'message': str(e)
        }), 500

@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle file too large error"""
    return jsonify({
        'error': 'File too large',
        'max_size': f'{MAX_FILE_SIZE / (1024 * 1024)}MB'
    }), 413

@app.errorhandler(404)
def not_found(error):
    """Handle 404 errors"""
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    logger.error(f"Internal server error: {str(error)}")
    return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('DEBUG', 'False').lower() == 'true'
    
    logger.info(f"Starting Gemini Image-to-Excel Service on port {port}")
    logger.info(f"Debug mode: {debug}")
    
    # Security: Only listen on localhost in production
    # Nginx will reverse proxy external requests
    app.run(
        host='127.0.0.1',  # Changed from 0.0.0.0 for security
        port=port,
        debug=debug
    )
