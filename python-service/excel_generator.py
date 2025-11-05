"""
Plain Excel Generator - Production Ready
Generates simple Excel files with no formatting
Outputs raw data only (no colors, borders, fonts, etc.)
"""

import os
import logging
from openpyxl import Workbook
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Generate plain Excel files from extracted table data.
    No styling, formatting, or visual enhancements.
    """
    
    def __init__(self):
        """Initialize Excel generator"""
        logger.info("✓ Plain Excel Generator initialized")
    
    def create_excel(self, tables, output_dir, filename=None):
        """
        Create a plain Excel file from extracted table data.
        
        Args:
            tables (list): List of table dictionaries with structure:
                [
                    {
                        'table_id': int,
                        'headers': [str, ...],
                        'data': [[str, ...], ...],
                        'metadata': {...}
                    }
                ]
            output_dir (str): Directory to save the Excel file
            filename (str, optional): Custom filename. If None, auto-generates timestamp-based name
        
        Returns:
            dict: Result with structure:
                {
                    'success': bool,
                    'excel_path': str,
                    'filename': str,
                    'total_tables': int,
                    'total_rows': int
                }
        """
        try:
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'extracted_table_{timestamp}.xlsx'
            
            # Ensure .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            excel_path = os.path.join(output_dir, filename)
            
            # Create workbook
            wb = Workbook()
            
            # Remove ALL default sheets to ensure clean workbook
            for sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            
            total_rows = 0
            
            # Process each table
            for table_idx, table in enumerate(tables, 1):
                # Create sheet for this table (always single sheet for Gemini extraction)
                sheet_name = "Extracted Table"  # Always use consistent name for single table
                ws = wb.create_sheet(title=sheet_name)
                
                # Get table data
                headers = table.get('headers', [])
                data = table.get('data', [])
                
                current_row = 1
                
                # Write headers (if available)
                if headers:
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col_idx, value=header)
                    current_row += 1
                
                # Write data rows
                for row_data in data:
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col_idx, value=cell_value)
                    current_row += 1
                    total_rows += 1
            
            # Save workbook
            wb.save(excel_path)
            
            file_size = os.path.getsize(excel_path)
            
            logger.info(f"✅ Plain Excel file created: {filename}")
            logger.info(f"   Tables: {len(tables)} | Rows: {total_rows}")
            logger.info(f"   Size: {file_size:,} bytes")
            
            return {
                'success': True,
                'file_path': excel_path,
                'excel_path': excel_path,
                'filename': filename,
                'total_tables': len(tables),
                'total_rows': total_rows,
                'file_size': file_size
            }
            
        except Exception as e:
            logger.error(f"Failed to create Excel file: {e}", exc_info=True)
            return {
                'success': False,
                'error': str(e),
                'excel_path': None,
                'filename': None,
                'total_tables': 0,
                'total_rows': 0
            }

