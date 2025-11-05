"""
Gemini 2.5 Flash - ENHANCED for Maximum Accuracy
Optimized prompt engineering for perfect handwriting OCR
"""

import google.generativeai as genai
from PIL import Image
import json
import os
from openpyxl import Workbook
from datetime import datetime

# Configure Gemini
GEMINI_API_KEY = "AIzaSyAe6-YcUO_JbNvPKudWRJe2X78dNpJhjFI"
genai.configure(api_key=GEMINI_API_KEY)

# Test image path
IMAGE_PATH = "E:/tool/IMG20251104124758.jpg"

print("=" * 80)
print("GEMINI 2.5 FLASH - ENHANCED HANDWRITTEN TABLE EXTRACTION")
print("=" * 80)
print(f"\nImage: {IMAGE_PATH}")
print(f"Model: gemini-2.5-flash (with enhanced prompting)")
print(f"\nProcessing with maximum accuracy settings...")
print("-" * 80)

# Load image
img = Image.open(IMAGE_PATH)
print(f"✓ Image loaded: {img.size[0]}x{img.size[1]} pixels")

# Initialize Gemini model with optimized settings
generation_config = {
    "temperature": 0.1,  # Low temperature for more deterministic/accurate output
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
}

model = genai.GenerativeModel(
    'gemini-2.5-flash',
    generation_config=generation_config
)

# ENHANCED prompt with better instructions
prompt = """You are an EXPERT OCR system with 99%+ accuracy on handwritten tables.

ANALYZE this handwritten table image and extract EVERY single detail with PERFECT accuracy.

CRITICAL INSTRUCTIONS:
1. CAREFULLY examine each cell - read handwriting character by character
2. For NUMBERS: Extract exact digits (e.g., "2250" not "2200")
3. For NAMES/TEXT: Preserve exact spelling even if unusual (e.g., "yaimly" stays "yaimly")
4. For MULTILINGUAL: Support English, numbers, and any other scripts
5. EMPTY CELLS: Mark as "" (empty string)
6. MERGED CELLS: If a cell spans multiple columns, put text in first cell
7. ROW DETECTION: Count ONLY actual data rows (skip header decorations)
8. COLUMN DETECTION: Identify exact number of columns based on vertical lines/alignment

STEP-BY-STEP PROCESS:
Step 1: Count vertical lines to determine exact number of COLUMNS
Step 2: Count horizontal lines/rows to determine exact number of ROWS
Step 3: Read EACH cell carefully from left to right, top to bottom
Step 4: For unclear handwriting, provide BEST interpretation (don't leave blank)
Step 5: Verify numbers add up correctly (if it's a calculation table)

OUTPUT FORMAT (STRICT JSON):
{
  "table_metadata": {
    "total_rows": <exact count>,
    "total_columns": <exact count>,
    "total_cells": <rows × columns>,
    "cells_with_content": <count>,
    "extraction_confidence": "high|medium|low",
    "table_type": "inventory|invoice|ledger|other"
  },
  "column_headers": ["Col1_Name", "Col2_Name", ...],
  "table_data": [
    ["row1_col1", "row1_col2", "row1_col3", ...],
    ["row2_col1", "row2_col2", "row2_col3", ...],
    ...
  ],
  "extraction_notes": "Any observations about handwriting quality, unclear cells, etc."
}

QUALITY CHECKS:
✓ Every cell must be examined (no skipping)
✓ Numbers must be exact (verify each digit)
✓ Text must preserve original spelling
✓ Row count must match actual data rows
✓ Column count must match vertical divisions

BEGIN EXTRACTION - Focus on ACCURACY over speed:"""

try:
    # Generate response with Gemini
    print("\n🤖 Gemini analyzing image with enhanced accuracy settings...")
    response = model.generate_content([prompt, img])
    
    print("\n" + "=" * 80)
    print("GEMINI ENHANCED RESPONSE")
    print("=" * 80)
    
    # Get response text
    response_text = response.text.strip()
    
    # Extract JSON from markdown code blocks
    if "```json" in response_text:
        json_start = response_text.find("```json") + 7
        json_end = response_text.find("```", json_start)
        json_text = response_text[json_start:json_end].strip()
    elif "```" in response_text:
        json_start = response_text.find("```") + 3
        json_end = response_text.find("```", json_start)
        json_text = response_text[json_start:json_end].strip()
    else:
        json_text = response_text
    
    # Parse JSON
    result = json.loads(json_text)
    
    # Display metadata
    meta = result['table_metadata']
    print(f"\n✓ Table Structure: {meta['total_rows']} rows × {meta['total_columns']} columns")
    print(f"✓ Total Cells: {meta['total_cells']}")
    print(f"✓ Cells with Content: {meta['cells_with_content']}")
    print(f"✓ Extraction Confidence: {meta['extraction_confidence']}")
    print(f"✓ Table Type: {meta['table_type']}")
    print(f"✓ Notes: {result.get('extraction_notes', 'N/A')}")
    
    print("\n" + "=" * 80)
    print("EXTRACTED TABLE DATA (Preview - First 10 rows)")
    print("=" * 80)
    
    # Show column headers
    if 'column_headers' in result:
        print(f"\nColumn Headers: {', '.join(result['column_headers'])}")
    
    # Show data preview
    for i, row in enumerate(result['table_data'][:10], 1):
        print(f"\nRow {i}:")
        for j, cell in enumerate(row, 1):
            if cell:  # Only show non-empty cells
                print(f"  Col {j}: '{cell}'")
    
    # Save JSON
    json_output_file = "e:/tool/convertnest-backend/python-service/outputs/gemini_enhanced_result.json"
    os.makedirs(os.path.dirname(json_output_file), exist_ok=True)
    with open(json_output_file, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)
    
    print(f"\n✓ JSON saved to: {json_output_file}")
    
    # ========================================================================
    # CREATE PLAIN EXCEL FILE (NO FORMATTING)
    # ========================================================================
    
    print("\n" + "=" * 80)
    print("GENERATING PLAIN EXCEL FILE (NO FORMATTING)")
    print("=" * 80)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Table"
    
    # Write column headers (plain text, no formatting)
    if 'column_headers' in result and result['column_headers']:
        for col_idx, header in enumerate(result['column_headers'], 1):
            ws.cell(row=1, column=col_idx, value=header)
    
    # Write data (plain text, no formatting)
    for row_idx, row_data in enumerate(result['table_data'], 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    # Save Excel file (plain, no formatting)
    excel_output_file = "e:/tool/convertnest-backend/python-service/outputs/gemini_plain_output.xlsx"
    wb.save(excel_output_file)
    
    print(f"\n✅ PLAIN EXCEL FILE CREATED!")
    print(f"   Path: {excel_output_file}")
    print(f"   Sheets: 1 (Extracted Table only)")
    print(f"   Rows: {meta['total_rows']} | Columns: {meta['total_columns']}")
    print(f"   Formatting: None (raw data only)")
    
    # File size
    file_size = os.path.getsize(excel_output_file)
    print(f"   Size: {file_size:,} bytes")
    
    # ========================================================================
    # COST ANALYSIS
    # ========================================================================
    
    print("\n" + "=" * 80)
    print("COST ANALYSIS")
    print("=" * 80)
    
    input_tokens = 774 + len(prompt.split()) * 1.3
    output_tokens = len(json_text.split()) * 1.3
    
    input_cost = (input_tokens / 1_000_000) * 0.30
    output_cost = (output_tokens / 1_000_000) * 2.50
    total_cost = input_cost + output_cost
    
    print(f"Input Tokens: ~{int(input_tokens)}")
    print(f"Output Tokens: ~{int(output_tokens)}")
    print(f"Cost per table: ${total_cost:.6f} (~{total_cost*100:.4f} cents)")
    print(f"Cost per 1,000 tables: ${total_cost * 1000:.2f}")
    print(f"Cost per year (12K tables): ${total_cost * 12000:.2f}")
    
    print("\n" + "=" * 80)
    print("✅ EXTRACTION COMPLETE - ALL FILES GENERATED!")
    print("=" * 80)
    print(f"\n📄 JSON: {json_output_file}")
    print(f"📊 EXCEL: {excel_output_file}")

except json.JSONDecodeError as e:
    print(f"\n❌ JSON Parse Error: {e}")
    print(f"\nRaw response:\n{response_text[:1000]}...")
    
    debug_file = "e:/tool/convertnest-backend/python-service/outputs/gemini_debug_enhanced.txt"
    with open(debug_file, 'w', encoding='utf-8') as f:
        f.write(response_text)
    print(f"\n✓ Debug output saved to: {debug_file}")

except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
print("TEST COMPLETE")
print("=" * 80)
